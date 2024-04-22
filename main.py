import os
import tkinter as tk
from tkinter import filedialog, messagebox

import openpyxl
import pandas as pd


class ExcelConverterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Text to Excel Converter")
        self.root.geometry("600x600")
        self.root.configure(bg="#0D567B")
        self.root.resizable(True, True)
        self.input_file_part_a = ""
        self.input_file_part_b = ""
        self.preprocessed_file = ""

        # DPHHS logo
        self.logo_path = "dphhs_logo.png"  # Adjust the path to your logo file
        self.logo_img = tk.PhotoImage(file=self.logo_path).subsample(5)  # Reduce the size by subsampling

        # Logo label
        self.logo_label = tk.Label(self.root, image=self.logo_img, bg="#0D567B")
        self.logo_label.place(relx=0.02, rely=0.02)  # Position the logo in the top-left corner
        # Button to select file for preprocessing
        self.select_preprocess_file_button = tk.Button(self.root,
                                                       text="Part 1: Select File for Preprocessing Part B’s Initial File",
                                                       command=self.select_preprocess_file,
                                                       bg="#1E88E5", fg="white", font=("Arial", 12))
        self.select_preprocess_file_button.pack(pady=5)
        # Button to select file for Part A
        self.select_file_part_a_button = tk.Button(self.root, text="Step 2: Select File for Part A",
                                                   command=self.select_file_part_a,
                                                   bg="#1E88E5", fg="white", font=("Arial", 12))
        self.select_file_part_a_button.pack(pady=5)

        # Button to select file for Part B
        self.select_file_part_b_button = tk.Button(self.root, text="Step 3: Select File for Part B",
                                                   command=self.select_file_part_b,
                                                   bg="#1E88E5", fg="white", font=("Arial", 12))
        self.select_file_part_b_button.pack(pady=5)

        # Button to process and save files
        self.process_button = tk.Button(self.root, text="Step 4: Process and Save All Selected Files", command=self.process_and_save,
                                        bg="#43A047", fg="white", font=("Arial", 12))
        self.process_button.pack(pady=5)

        # Label to show preprocessed file
        self.preprocessed_label = tk.Label(self.root, text="", bg="#0D567B", fg="white", font=("Arial", 12))
        self.preprocessed_label.pack(pady=5)

    def select_file_part_a(self):
        files = filedialog.askopenfilenames(filetypes=[("Text files", "*.txt")])
        if files:
            self.input_file_part_a = files
            messagebox.showinfo("File Selected", "File(s) selected for Part A.")

    def select_file_part_b(self):
        file = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
        if file:
            self.input_file_part_b = file
            messagebox.showinfo("File Selected", "File selected for Part B.")

    def preprocess_file(self):
        input_file = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
        if input_file:
            # Prompt user to select destination folder and filename for preprocessed file
            output_file = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt")])
            if output_file:
                success = clean_file(input_file, output_file)
                if success:
                    self.preprocessed_file = output_file
                    # Update this line to output the file path to a message box instead
                    messagebox.showinfo("Preprocessing Completed", f"Preprocessed file saved as: {output_file}")
                else:
                    messagebox.showerror("Error", "Error occurred during preprocessing.")

    def select_preprocess_file(self):
        input_file = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
        if input_file:
            # Prompt user to select destination folder and filename for preprocessed file
            output_file = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt")])
            if output_file:
                success = clean_file(input_file, output_file)
                if success:
                    self.preprocessed_file = output_file
                    # Update this line to output the file path to a message box instead
                    messagebox.showinfo("Preprocessing Completed", f"Preprocessed file saved as: {output_file}")
                else:
                    messagebox.showerror("Error", "Error occurred during preprocessing.")

    def process_and_save(self):
        if not self.input_file_part_a or not self.input_file_part_b:
            messagebox.showwarning("Missing File(s)", "Please select file(s) for both Part A and Part B.")
            return
        if not self.preprocessed_file:
            messagebox.showwarning("Missing Preprocessed File", "Please preprocess a file first.")
            return
        try:
            # Prompt user to select destination folder for Part A output
            output_folder_a = filedialog.askdirectory(title="Select Output Folder for Part A")
            if not output_folder_a:
                return
            # Process and save Part A files
            for idx, input_file in enumerate(self.input_file_part_a):
                output_file_a = os.path.join(output_folder_a, f"output_part_a_{idx + 1}.xlsx")
                self.process_and_save_part_a(input_file, output_file_a)

            # Prompt user to select destination folder for Part B output
            output_folder_b = filedialog.askdirectory(title="Select Output Folder for Part B")
            if not output_folder_b:
                return
            # Process and save Part B file
            output_file_b = os.path.join(output_folder_b, "output_part_b.xlsx")
            self.process_and_save_part_b(self.input_file_part_b, output_file_b)
            messagebox.showinfo("Processing Completed", "Files processed successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def process_and_save_part_a(self, source_file, destination_file):
        if not self.input_file_part_a:
            messagebox.showwarning("No File", "Please select TXT files first.")
            return

        # Open the input text file and create a new Excel workbook
        with open(source_file, 'r', encoding='utf-8', errors='ignore') as infile:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Data'

            # Define the headers
            headers = [
                'First Name', 'Last Name', 'Social Security', 'CAPS ID', 'Date of Birth',
                'Service Code', 'Payment Begin', 'Payment End', 'Post Date', 'County',
                'Payment #', 'Seq #', 'Adj Seq #', 'Units', 'Payment Amt', 'Funding Source',
                'Fund Adj Sequence #', 'Speed Chart', 'Provider', 'Facility', 'Provider Name',
                'Overpayment', 'Recoup', 'Pay Adjust', 'Sabhrs Form', 'SABHRS Doc #',
                'SSN DBKEY', 'Pay Post Date'
            ]

            # Write the header row to the Excel sheet
            sheet.append(headers)

            # Loop through each line in the input file, skipping the first line
            lines = infile.readlines()[1:]
            total_lines = len(lines)
            for line_idx, line in enumerate(lines):
                if 'Superior' in line[40:49]:
                    # Skip lines with 'Superior' in positions 41-49
                    continue

                # Clean the line by removing non-printable characters
                cleaned_line = ''.join(char for char in line if char.isprintable())

                self.root.update_idletasks()

                # Split the cleaned line into fields based on fixed positions
                data = [
                    # Column headers:
                    # 'First Name'
                    cleaned_line[29:37].strip().upper(),
                    # 'Last Name'
                    cleaned_line[17:28].strip().upper(),
                    # 'Social Security'
                    cleaned_line[8:17].strip(),
                    # 'CAPS ID'
                    cleaned_line[0:8].strip(),
                    # 'Date of Birth'
                    cleaned_line[40:51].strip(),
                    # 'Service Code'
                    cleaned_line[51:56].strip(),
                    # 'Payment Begin'
                    cleaned_line[56:66].strip(),
                    # 'Payment End'
                    cleaned_line[66:76].strip(),
                    # 'Post Date'
                    cleaned_line[76:86].strip(),
                    # 'County'
                    cleaned_line[86:89].strip(),
                    # 'Payment #'
                    cleaned_line[89:98].strip(),
                    # 'Seq #'
                    cleaned_line[98:101].strip(),
                    # 'Adj Seq #'
                    cleaned_line[101:103].strip(),
                    # 'Units'
                    cleaned_line[103:109].strip(),
                    # 'Payment Amt'
                    cleaned_line[110:120].strip(),  # This is where the Payment Amt field is extracted
                    # 'Funding Source'
                    cleaned_line[120:123].strip(),
                    # 'Fund Adj Sequence #'
                    cleaned_line[123:125].strip(),
                    # 'Speed Chart'
                    cleaned_line[125:130].strip(),
                    # 'Provider'
                    cleaned_line[130:137].strip().title(),
                    # 'Facility'
                    cleaned_line[137:140].strip(),
                    # 'Provider Name'
                    cleaned_line[140:155].strip(),
                    # 'Overpayment'
                    cleaned_line[155:156].strip(),
                    # 'Recoup'
                    cleaned_line[156:157].strip(),
                    # 'Pay Adjust'
                    cleaned_line[157:158].strip(),
                    # 'Sabhrs Form'
                    cleaned_line[158:161].strip(),
                    # 'SABHRS Doc #'
                    cleaned_line[161:169].strip(),
                    # 'SSN DBKEY'
                    'NULL',
                    # 'Pay Post Date'
                    cleaned_line[76:86].strip()
                ]

                # Check if Payment Amt contains non-numeric characters (excluding '-')
                payment_amt = data[14].replace(',', '')  # Remove commas
                if payment_amt.replace('.', '').lstrip('-').isdigit():
                    # Check if it's numeric
                    data[14] = float(payment_amt)  # Convert to float if it's numeric
                else:
                    data[14] = 0.0  # Set to 0.0 if non-numeric or handle based on requirements

                # Write the fields to the Excel sheet
                if not data[4].__contains__('SUPERIOR'):
                    # Write the fields to the Excel sheet
                    sheet.append(data)

            # Save the Excel workbook
            workbook.save(destination_file)

            # Apply numeric format to the 'Payment Amt' column (column O)
            # Open the workbook with openpyxl
            wb = openpyxl.load_workbook(destination_file)
            # Select the active sheet


        ws = wb.active
        # Apply the numeric format to the 'Payment Amt' column (column O)
        for row in ws.iter_rows(min_row=2, min_col=15, max_col=15, max_row=ws.max_row):
            for cell in row:
                cell.number_format = '#,##0.00'

        # Save the workbook
        wb.save(destination_file)

    def process_and_save_part_b(self, source_file, destination_file):
        try:
            # Read text data from the source file
            with open(source_file, 'r') as file:
                text_data = file.read()
            # Split text into lines
            lines = text_data.strip().split('\n')
            # Initialize lists to store data
            regions = []
            region_names = []
            county_nums = []
            county_names = []
            first_names = []
            last_names = []
            current_worker_ids = []
            error_codes = []
            error_types = []
            payment_numbers = []
            line_items = []
            client_ids = []
            facilities = []
            service_codes = []
            begin_dates = []
            end_dates = []
            error_dates = []
            entry_amounts = []
            # Process each line of text
            current_region = None
            current_region_name = None
            current_first_name = None
            current_last_name = None
            current_worker_id = None
            current_county_num = None
            current_county_name = None
            skip_lines = False
            for line in lines:
                if "REGION LINE" in line:
                    current_region = line[30:32]
                    current_region_name = line[39:52]
                elif "COUNTY LINE" in line:
                    current_county_num = line[30:33]
                    current_county_name = line[39:66]
                elif "Assigned Worker LINE" in line:
                    current_first_name = line[69:76]
                    current_last_name = line[48:59]
                    current_worker_id = line[39:45]  # Assign current worker ID here
                elif "Entry Line 3" in line:
                    regions.append(current_region)
                    region_names.append(current_region_name)
                    county_nums.append(current_county_num)
                    county_names.append(current_county_name)
                    first_names.append(current_first_name)
                    last_names.append(current_last_name)
                    current_worker_ids.append(current_worker_id)  # Append to list here
                    error_codes.append(line[14:18])
                    error_types.append(line[20:25])
                    payment_numbers.append(line[27:36])
                    line_items.append(line[38:41])
                    # Client ID
                    client_ids.append(line[44:52])
                    facilities.append(line[54:65])
                    service_codes.append(line[85:90])
                    begin_dates.append(line[110:121])
                    end_dates.append(line[122:133])
                    error_dates.append(line[134:145])
                elif "Entry Line 6" in line:
                    entry_amounts.append(line[31:39])
                # Check for STATE OF MONTANA pattern to skip lines
                if "STATE OF MONTANA" in line:
                    skip_lines = True
                elif skip_lines:
                    if "RUN DATE" in line:
                        skip_lines = False
            # Remove commas from entry amounts and then convert to numeric
            entry_amounts = [amount.replace(',', '') for amount in entry_amounts]
            entry_amounts = pd.to_numeric(entry_amounts, errors='coerce')  # Handle invalid entries
            # Create a DataFrame using pandas
            df = pd.DataFrame({
                'ERROR CODE': error_codes,
                'ERROR TYPE': error_types,
                'PAYMENT #': payment_numbers,
                'LINE ITEM': line_items,
                'CLIENT ID': client_ids,
                'FACILITY #': facilities,
                'SERVICE CODE': service_codes,
                'BEGIN DATE': begin_dates,
                'END DATE': end_dates,
                'ERROR DATE': error_dates,
                'WORKER ID': current_worker_ids,  # Use current_worker_ids list
                'WORKER LAST NAME': last_names,
                'WORKER FIRST NAME': first_names,
                'COUNTY #': county_nums,
                'COUNTY NAME': county_names,
                'REGION': regions,
                'SERVICE AMOUNT': entry_amounts,
            })
            # Create an ExcelWriter object with xlsxwriter engine
            with pd.ExcelWriter(destination_file, engine='xlsxwriter') as writer:
                # Write DataFrame to Excel using the specified writer
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                # Get the workbook and worksheet objects
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']
                # Define a numeric format for the 'Service Amount' column
                numeric_format = workbook.add_format({'num_format': '#,##0.00'})
                # Apply the numeric format to the 'Service Amount' column
                worksheet.set_column('Q:Q', None, numeric_format)
            return True
        except Exception as e:
            print(e)
            return False

    def drop_handler(self, event):
        # Get the list of dropped files
        files = event.data
        if files:
            # Take the first file dropped
            preprocessed_file = files[0]
            self.preprocessed_file = preprocessed_file
            self.preprocessed_label.config(text=f"Preprocessed file: {preprocessed_file}")
            messagebox.showinfo("File Selected", "Preprocessed file selected for Part B processing.")


def clean_file(input_file, output_file):
    try:
        # Open the input file in read mode and output file in write mode
        with open(input_file, 'r') as f_in, open(output_file, 'w') as f_out:
            # Flag to track whether to skip lines
            skip_lines = False
            # Counter for lines within each entry
            line_counter = 0
            # Flag to track the start of a new entry
            start_of_entry = True
            # Read each line in the input file
            for line in f_in:
                # Check for REGION, COUNTY, ASSIGNED, and ERROR patterns
                if "REGION" in line:
                    f_out.write("REGION LINE: " + line)
                    start_of_entry = False
                    continue
                if "COUNTY" in line:
                    f_out.write("COUNTY LINE: " + line)
                    start_of_entry = False
                    continue
                if "ASSIGNED" in line:
                    f_out.write("Assigned Worker LINE: " + line)
                    start_of_entry = False
                    continue
                if "ERROR" in line:
                    line_counter = 0  # Reset line counter for new entry
                    f_out.write(f"Entry Line {line_counter + 1}: {line}")
                    start_of_entry = False
                    continue

                # Check if the line contains the unwanted pattern to start skipping lines
                if "STATE OF MONTANA" in line:
                    skip_lines = True
                    start_of_entry = False
                    continue  # Skip this line

                # Check if we are in the middle of the unwanted section
                if skip_lines:
                    # Check if we've reached the end of the unwanted section
                    if "RUN DATE" in line:
                        skip_lines = False
                    start_of_entry = False
                    continue  # Skip this line

                # Check if the line is empty (contains only whitespace characters)
                if line.strip():  # If the line is not empty
                    # Write the line to the output file
                    if start_of_entry:
                        line_counter = 0  # Start counting lines within entry from 0
                        f_out.write(f"Entry Line {line_counter + 1}: {line}")
                        start_of_entry = False
                    else:
                        line_counter += 1
                        f_out.write(f"Entry Line {line_counter + 1}: {line}")
                    start_of_entry = False

        return True
    except Exception as e:
        print(e)
        return False


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelConverterApp(root)
    root.mainloop()
